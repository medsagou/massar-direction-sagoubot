import openpyxl

# Load the .xlsm workbook
workbook_path = 'pyTable.xlsm'  # Replace with your Excel file path
wb = openpyxl.load_workbook(workbook_path, keep_vba=True)  # keep_vba=True is needed for .xlsm files

# Text to be added in F4 and G4
text_f4 = "رقم هاتف الأب"
text_g4 = "رقم هاتف الأم"

# Iterate through all sheets in the workbook
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Clear data in F4:F50 and G4:G50
    for row in range(10, 70):  # Rows from 4 to 50
        ws[f'F{row}'] = None
        ws[f'G{row}'] = None

    # Add text to F4 and G4
    ws['F10'] = text_f4
    ws['G10'] = text_g4

# Save the changes to the workbook
wb.save('modified_' + workbook_path)  # Save with a new name or overwrite the original file
