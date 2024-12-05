import pandas as pd
import openpyxl
from utilities import print_error, print_success, print_info
from utilities import check_exist_file
import xlrd
import re
import os
import time
class Read_Db:
    def __init__(self, input_file = r"data_to_manage/file_data.xls", template_file = "data_to_manage/template.xlsx", output_file = r"C:\Users\HP\Desktop\abs2025\absence.xlsx", df = "", required_classes=[], progress_bar="", console=""):
        self.index = {0: "CLASS_StudentIndex",
                      1: "Niveau",
                      2: "class_name",
                      3: "student_index",
                      "Unnamed: 1": "CNE",
                      "Unnamed: 2": "nom",
                      "Unnamed: 3": "prenom"}
        self.input_file = input_file
        self.output_file = output_file
        self.template_file = template_file
        self.df = df
        self.init_cell = ["A"]
        self.start_col = 'A'
        self.end_col = 'C'
        # self.workbook_output = self.get_workbook(output_file)
        self.workbook_output = ""
        self.required_classes = required_classes
        self.progress_bar = progress_bar
        self.console = console

    def get_key(self, val):
        for key, value in self.index.items():
            if val == value:
                return key
        return "key doesn't exist"

    def get_data_from_xlsb(self):
        xlsb_file = pd.ExcelFile(self.input_file)
        df = xlsb_file.parse('Feuil3', header=None)  #
        self.df = df
        return df
    def get_df_from_xls(self):
        xls = pd.ExcelFile(self.input_file)
        workbook = self.get_data_from_xls()
        sheet_names = xls.sheet_names
        data = {}
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            df = pd.read_excel(self.input_file, sheet_name=sheet_name)
            class_name = sheet.cell_value(7, 2)
            data[class_name] = df
        self.df = data
        return data

    def get_data_from_xls(self):        # new data function
        return xlrd.open_workbook(self.input_file)
    def get_classes_name_from_xls(self):
        workbook = self.get_data_from_xls()
        classes = []
        sheet_names = workbook.sheet_names()
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            class_name = sheet.cell_value(7, 2)
            # print(class_name)
            classes.append(class_name)
        return classes

    def get_workbook(self, file_name):
        print(file_name)
        workbook = openpyxl.load_workbook(file_name)
        return workbook


    def get_workbook_sheet(self, workbook ,sheet):
        return workbook[sheet]

    def add_value_to_sheet(self, worksheet, cell, value):
        cell_to_update = worksheet[cell]
        cell_to_update.value = value
        return


    def create_copy_sheet(self, class_name = "", workbook = "", source_sheet = ""):
        new_sheet = workbook.copy_worksheet(source_sheet)
        new_sheet.title = class_name
        new_sheet.sheet_view.rightToLeft = True
        return


    def get_column_list_from_df(self, column_key):
        if self.df == "":
            self.get_df_from_xls()

        L = list(set(self.df.values[:, column_key].tolist()))
        try:
            L.remove("0")
        except ValueError:
            pass
        try:
            L.remove(0)
        except ValueError:
            pass
        return L
    def restart_workbook_output(self):
        self.workbook_output.close()
        self.workbook_output = self.get_workbook(self.output_file)
        return
    def get_sheet_names_workbout_output(self):
        self.workbook_output = self.get_workbook(self.output_file)
        return self.workbook_output.sheetnames




    def create_all_class_sheet(self):
        if check_exist_file(self.output_file):
            # class_in_sheet = self.get_sheet_names_workbout_output()
            # with open(self.output_file, 'w') as f:
            #     f.close()
            os.remove(self.output_file)
            print_info("WE REMOVED THE OUTPUT FILE TO CREATE NEW ONE", console=self.console)
        # else:
        #     class_in_sheet = []
        # classes_list = self.get_column_list_from_df(column_key=self.get_key("class_name"))

        workbook = openpyxl.load_workbook(self.template_file)
        source_sheet = workbook["BaseSheet"]
        classes_list = self.get_classes_name_from_xls()
        print(len(classes_list))
        # print(classes_list)

        # print(self.required_classes)
        i = 0
        for classe in classes_list:
            # if classe in class_in_sheet:
            #     print_error(f"SHEET FOR {classe} ALREADY EXIST")
            #     continue
            # if not in college just skipit
            # print(classe)
            # print(classe.split("-")[0][1:])
            # pattern = re.compile(rf"^(\d*)({'|'.join(map(re.escape, self.required_classes))})")
            # if not bool(pattern.match(classe)):
            #     continue
            if 'ASCPEB' in classe or 'APIC' in classe or 'ASCG' in classe:
                continue
            i += 1
            print_info(f"CREATE A SHEET FOR {classe} CLASS", console=self.console)
            if classe != "":
                self.create_copy_sheet(class_name=classe, workbook=workbook, source_sheet = source_sheet)
        print_info(f"{i} Classes", console=self.console)

        workbook.save(str(self.output_file))
        workbook.close()
        return

    def fill_all_class_sheets(self):
        self.create_all_class_sheet()
        # already check above
        if str(self.df) == "":
            print_info("GETTING THE DATA...", console=self.console)
            self.get_data_from_xls()
        # print_info("RESTARTING WORKSHEET")
        # self.restart_workbook_output()
        self.workbook_output = self.get_workbook(self.output_file)
        class_in_sheet = list(self.get_sheet_names_workbout_output())
        # print(class_in_sheet)
        for k in range(len(class_in_sheet)):
            # print(f"{k+1}/{len(class_in_sheet)}")
            self.progress_bar.set((k+1)/len(class_in_sheet))
            worksheet = self.get_workbook_sheet(workbook = self.workbook_output, sheet=class_in_sheet[k])
            i = 0
            print_info(f"WORKING ON {class_in_sheet[k]} CLASS DATA TO SHEET", console=self.console)
            # column = db.df["3ASCG-5"].columns.tolist()
            #
            # for index, row in db.df["3ASCG-5"].iterrows():
            #     if pd.isna(row[column[23]]):
            #         continue
            #     print(row[column[23]], row[column[16]], row[column[12]])
            index_student = 0
            self.get_df_from_xls()
            if class_in_sheet[k] == 'BaseSheet':
                continue
            # for index, row in self.df[class_in_sheet[k]].iterrows():
            #     print(index, row)
            # time.sleep(20000)
            for index, row in self.df[class_in_sheet[k]].iterrows():
                if pd.isna(row[self.get_key("CNE")]):
                    continue
                if index_student == 0:
                    index_student += 1
                    continue
                i += 1
                    # print(row)
                for col in range(ord(self.start_col), ord(self.end_col) + 1):
                    if chr(col) == "A":
                        self.add_value_to_sheet(worksheet=worksheet, cell=chr(col) + str(9 + i), value=index_student)
                    elif chr(col) == "B":
                        self.add_value_to_sheet(worksheet=worksheet, cell=chr(col) + str(9 + i), value=row[self.get_key("CNE")])
                    elif chr(col) == "C":
                        self.add_value_to_sheet(worksheet=worksheet, cell=chr(col) + str(9 + i),
                                                value=str(row[self.get_key("prenom")] + " " + str(row[self.get_key("nom")])))
                    self.add_value_to_sheet(worksheet=worksheet, cell="BA" + str(9 + i), value=str(row[self.get_key("prenom")] + " " + str(row[self.get_key("nom")])))
                    if i > 49:
                        return

                index_student += 1


            # add number of students
            self.add_value_to_sheet(worksheet=worksheet, cell="AO6", value=str(i))
            # add class name
            self.add_value_to_sheet(worksheet=worksheet, cell="D6", value=class_in_sheet[k])
            self.workbook_output.save(self.output_file)
            # self.workbook_output.close()
        print_success("Your lists is generated successfully", console=self.console)
        print_success(f"Your file path:  {self.output_file}", console=self.console)
        return



#
# db = Read_Db()
# db.get_df_from_xls()
# column = db.df["3ASCG-5"].columns.tolist()
# # print(column[23], column[16], column[12])
# # # Unnamed: 23 Unnamed: 16 Unnamed: 12
# for index, row in db.df["3ASCG-5"].iterrows():
#     if pd.isna(row[column[23]]):
#         continue
#     print(row[column[23]], row[column[16]], row[column[12]])

# for i in range(len(column)):
#     print(db.df['3ASCG-5'][column[i]])
# data = db.fill_all_class_sheets()
# db.create_all_class_sheet()
# print(db.get_df_from_xls().keys())
# add_value_to_sheet(worksheet=db.workbook_output["3ASCG-2"], cell="B11", value = "test")
# db.workbook_output.save("db/Book13.xlsx")
# db.workbook_output.close()
# # db.get_data_from_xlsb()
# # print(db.get_classes_list_from_df())
#
# db.fill_all_class_sheets()


